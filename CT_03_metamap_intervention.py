#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import print_function

import io
import re
import sys
import time
import logging
if sys.version_info[0] < 3:
    reload(sys)
    sys.setdefaultencoding('utf-8')

#make utils folder visible to system for imports
sys.path.insert(1,"utils")
import csv
import openpyxl
from openpyxl import load_workbook
from lxml import etree

import pickle
import unicodedata
import subprocess
import multiprocessing
from timeout import timeout
from prohibited_words import p_words

import spacy
import string

csv.field_size_limit(58000000)

logging.basicConfig(filename=sys.argv[4],level=logging.DEBUG)
# delimiter used in CSV files
MMQUERY_CSV_DELIMITER = '\t'
FAILED_REQUESTS = dict()
# default field value in CSV files
# this is currently '0' due to Neo4j requirements
MMQUERY_CSV_DEFAULT_FIELD = '0'

# Excel file containing the UMLS category codes
MMQUERY_UMLS_DICT = 'umls_statex.xlsx'
MMQUERY_UMLS_SEMTYPES_ID = 'sem_types'
MMQUERY_UMLS_SEMGROUPS_ID = 'sem_groups'
MMQUERY_UMLS_STRING_SEP = '|'

REGEX = "[+-]?(([0-9]*[.])?[0-9]+(\s*[-]+\s*))?([0-9]*[.])?[0-9]"

# set of concepts which have its UMLS priority and MetaMap confidence modified
MMQUERY_SCORE_MOD = {}
MMQUERY_SCORE_MOD['Symptoms'] = [2.9, None]
MMQUERY_SCORE_MOD['Proteins'] = [None, 100]
MMQUERY_SCORE_MOD['Knock-out'] = [None, 100]
MMQUERY_SCORE_MOD['Exacerbation'] = [None, 100]

# function that builds a MetaMap query given a list of newline separated terms
MMQUERY_MMSH = '/tools/metamap_2020/public_mm/bin/metamap -Z 2023AB -V custom2025AB -Cz -R AIR,AOD,AOT,ATC,CCS,CCSR_ICD10PCS,CHV,COSTAR,CSP,CVX,DRUGBANK,DXP,FMA,GO,HCPCS,HGNC,HL7V2.5,HL7V3.0,HPO,ICD10CM,ICD10PCS,ICD9CM,ICF,ICF-CY,ICPC,LCH,LCH_NW,LNC,MCM,MDR,MED-RT,MEDLINEPLUS,MSH,MTH,MTHCMSFRF,MTHICD9,MTHMST,MTHSPL,MVX,NCBI,NCI,OMIM,PDQ,QMR,RAM,RXNORM,SNOMEDCT_US,SNOMEDCT_VET,SOP,SPN,SRC,USP,USPMG,UWDA,VANDF  --sldi --prune 15 --XMLf'
MMQUERY_CMD = lambda s: 'echo "' + s + '" | ' + MMQUERY_MMSH

BLACKLISTED_CUIS = dict()
wb =  load_workbook(filename = "utils/Universal_statex_blacklist.xlsx")
sheet = wb['universal_statex_blacklist']
start_row = 2
end_row = sheet.max_row+1
for row in range(start_row, end_row):
    concept = sheet['A'+str(row)].value
    cui = sheet['B'+str(row)].value.upper().strip()
    BLACKLISTED_CUIS[cui] = concept

ALLOWED_CONCEPT_CATS = ['CHEM', 'T168', 'LIVB', 'PROC', 'T074', 'T203','T025']
# set of words that are removed in order for MetaMap to match items
MMQUERY_WORD_BLACKLIST = ['impact', 'accumulation', 'program']

# set of UMLS categories that cannot be the result of a MetaMap query
MMQUERY_SEMTYPE_BLACKLIST = ['CONC|Concepts & Ideas|T082|Spatial Concept']


def read_umls_catcodes():
    """
    Read UMLS categories, together with their abbreviations and priorities
        Outputs:
            - abbrv2cat: dictionary from abbreviations to category names
            - cat2group: dictionary from category names to its group at priority
    """
    workbook = openpyxl.load_workbook(filename = MMQUERY_UMLS_DICT)

    # obtain UMLS semantic categories with their abbreviations
    abbrv2cat = dict()
    for row in workbook[MMQUERY_UMLS_SEMTYPES_ID].iter_rows(min_row=2, max_col=1):
        for cell in row:
            cat_forms = cell.value.split(MMQUERY_UMLS_STRING_SEP)
            if cat_forms[0] not in abbrv2cat:
                abbrv2cat[cat_forms[0]] = cat_forms[2]

    # obtain semantic groups with their priorities
    cat2group = dict()
    for row in workbook[MMQUERY_UMLS_SEMGROUPS_ID].iter_rows(min_row=2, max_col=2):
        for cell_pair in zip(row[:-1], row[1:]):
            cat_group = cell_pair[0].value.split(MMQUERY_UMLS_STRING_SEP)
            try:
                cat_priority = float(cell_pair[1].value)
            except ValueError:
                cat_priority = 1
                raise Exception(MMQUERY_UMLS_DICT + ': the priority of a UMLS category must be a number.')
            if cat_group[-1] not in cat2group:
                cat2group[cat_group[-1]] = [cat_group, cat_priority]

    return abbrv2cat, cat2group


def substract_hyphens(text):
    """
    Substract hyphens from the input text under certain circumstances
        Inputs:
            - text: string
        Outputs:
            input text with some hyphens removed
    """
    tokens = text.split()
    for i in range(len(tokens)):
        split_token = tokens[i].split('-')
        if len(split_token) == 2 and split_token[0].isdigit() and len(split_token[1]) < 3:
            tokens[i] = tokens[i].replace('-', '')
    return ' '.join(tokens).replace(MMQUERY_CSV_DELIMITER, '')


def map_global_terms(istring):
    """
    Replaces the input string with a global term if defined
        Inputs:
            - istring: string
        Outputs:
            global term corresponding to the string or the itself
    """
    if istring in global_text2text:
        return u''+global_text2text[istring]
    return u''+istring


def normalize_char(char):
    """
    Normalize an input character
        Inputs:
            - char: character
        Outputs:
            normalized form of the input character (e.g. 'α' -> 'alpha')
    """
    if char.encode('utf-8') in sym2text:
        return sym2text[char.encode('utf-8')]
    return char


def clean_entity_string(request_string):
    """
    Clean an input string so that it can match UMLS entities
        Inputs:
            - request_string: string
        Outputs:
            cleaned form of the input string
    """
    request_string_tokens = request_string.split()
    request_string_filtered = ''

    for i in range(len(request_string_tokens)):
        if request_string_filtered:
            request_string_filtered += ' '
        if request_string_tokens[i].lower() not in MMQUERY_WORD_BLACKLIST:
            token_lemma = ''.join(x.lemma_ for x in MMQUERY_NLP(u'' + request_string_tokens[i]))
            if token_lemma not in MMQUERY_WORD_BLACKLIST:
                request_string_filtered += request_string_tokens[i]

    # take care of greek letters and other symbols
    request_string_filtered = u'' + request_string_filtered
    request_string_clean =  ''.join([normalize_char(x) for x in request_string_filtered])
    return request_string_clean


def clean_query_string(request_string):
    """
    Clean an input string so that it can match be employed as part of a MetaMap query
        Inputs:
            - request_string: string
        Outputs:
            cleaned form of the input string
    """
    plain_string = unicodedata.normalize('NFKD', u'' + request_string).encode('ASCII', 'ignore')
    return ''.join(c if c not in MMQUERY_FORBIDDEN_SYMBOLS else ' ' for c in plain_string)


def update_record(best_rating, current_rating, best_record, current_record):
    """
    Decides whether a MetaMap candidate concept is better than the current best and potentially replaces it
        Inputs:
            - best_rating: scores of the best candidate as [priority, confidence]
            - current_rating: scores of the current candidate as [priority, confidence]
            - best_record: best MetaMap candidate
            - current_record: current MetaMap candidate
        Outputs:
            the best record after updating together with its rating
    """
    # update the best record if the current record has a higher priority
    if (current_rating[0] > best_rating[0]) or (current_rating[0] == best_rating[0] and current_rating[1] > best_rating[1]):
        current_record_semtypes = current_record[3]
        for black_semtype in MMQUERY_SEMTYPE_BLACKLIST:
            current_record_semtypes = current_record_semtypes.replace(black_semtype, '')
        # update only when the UMLS category is not blacklisted
        if current_record_semtypes:
            best_rating = current_rating
            best_record = current_record
    return best_record, best_rating


def heuristic_fixes(request_string):

    """
    String manipulations before metamap invocation
    """

    unigrams = {"glaucoma":"glaucomas","anemia":"blood anemia", "anaemia":"blood anemia",
                "hiv":"human immunodeficiency virus", "ms":"multiple sclerosis",
                "hyperlipademia":"hyperlipidemia", "mets":"metabolic syndrome",
                "rebleeding":"recurrent hemorrhage","tb":"tuberculosis","spp":"species",
                "spp.":"species", "tumor":"neoplasm","mtx":"methotrexate","MTX":"methotrexate"}

    ngrams = { "vitamin d insufficiency":"vitamin d deficiency", "alzheimer 's disease":"alzheimer",
                "alzheimer's disease":"alzheimer", "alzheimer disease":"alzheimer","alzheimers disease":"alzheimer",
                "myocardial infarction":"heart attack","cigarette smoking":"tobacco smoking behavior",
                "high blood pressure":"blood pressure, high","spp .":"species", "huntington 's disease":"huntington chorea",
                "huntington's disease":"huntington chorea", "huntington disease":"huntington chorea",
                "huntingtons disease":"huntington chorea","radiation therapy":"therapeutic radiology procedure",
                "radiotherapy":"therapeutic radiology procedure","surgical procedure":"operative surgical procedures",
                "pain therapy":"pain management"}



    tokens = request_string.split()
    for idx,tok in enumerate(tokens):
        if tok in unigrams:
            tokens[idx] = unigrams[tok]

    request_string_clean = " ".join(tokens)

    for token,manipulation in ngrams.items():
        request_string_clean = request_string_clean.replace(token,manipulation)

    regex_patterns = ['\xe2\x84\xa2','\xc2\xa9','\xc2\xae', REGEX+'+\s*g/kg', REGEX+'+\s*mg/m2/day',REGEX+'+\s*mg/m\^2/day',
                    REGEX+'+\s*mg/m2',REGEX+'+\s*mg/m\^2',REGEX+'+\s*mg/kg',REGEX+'+\s*mg/day',REGEX+'+\s*mcg/kg',
                    REGEX+'+\s*mg/ml',REGEX+'+\s*mg/d',REGEX+'+\s*mg']

    for pattern in regex_patterns:
        request_string_clean = re.sub(pattern,"",request_string_clean)

    tokens = request_string_clean.split()
    permited_tokens = []
    for tok in tokens:
        if tok not in p_words:
            permited_tokens.append(tok)


    request_string_clean = " ".join(permited_tokens)
    request_string_clean = request_string_clean.replace("()","")


    return request_string_clean


def get_best_record(xml, request_string, umls_cat_abbrv, umls_cat_group):
    """
    Obtains the best and the second best candidate concepts given a MetaMap response
        Inputs:
            - xml: metamap response to a single concept query as a formatted XML file
            - request_string: string representing a single concept
            - umls_cat_abbrv: dictionary from abbreviations to category names
            - umls_cat_group: dictionary from category names to its group at priority
        Outputs:
            - best_candidate_record: best ([confidence, cui, preferred_form, semtype, matched_form], all_concepts) tuple
    """
    # obtain the root of the XML file
    try:
        root = etree.fromstring(xml)
    except:
        print("XML format error")
        logging.warning('XML format error on request: ' + str(request_string))
        FAILED_REQUESTS[request_string] = 1

        return [MMQUERY_CSV_DEFAULT_FIELD]*5,[]

    all_concepts = []
    candidates = []

    mapcands = root.findall('.//MappingCandidates')
    for mapcand in mapcands:
        if mapcand != None and len(mapcand):
            for concept in mapcand:
                # information fields for each concept provided by MetaMap
                concept_score = 0
                concept_cui = ''
                concept_candidate_matched = ''
                concept_candidate_preferred = ''
                # matching word positions in the request string
                concept_matched_spans = []
                # UMLS categories corresponding to the matched concept
                concept_semtypes = []
                concept_semtype_priority = 0


                concept_score = -int(concept.find('CandidateScore').text)
                concept_cui = concept.find('CandidateCUI').text
                concept_candidate_matched = concept.find('CandidateMatched').text.replace('*', '').replace('^', '')
                concept_candidate_preferred = concept.find('CandidatePreferred').text.replace('*', '').replace('^', '')
                for mspan in concept.find('.//MatchMaps'):
                    mspan_start_index = int(mspan.find('TextMatchStart').text)-1
                    mspan_end_index = int(mspan.find('TextMatchEnd').text)-1
                    concept_matched_spans.append([mspan_start_index, mspan_end_index])

                # switch between preferred and matched candidates when the former is empty
                if not concept_candidate_preferred:
                    concept_candidate_preferred = concept_candidate_matched
                    concept_candidate_matched = MMQUERY_CSV_DEFAULT_FIELD

                # find the highest priority within the UMLS categories
                for semtype in concept.find('.//SemTypes'):
                    semtype_group = umls_cat_group[umls_cat_abbrv[semtype.text]]
                    semtype_category = semtype_group[0]
                    semtype_priority = semtype_group[1]
                    concept_semtypes.append(MMQUERY_UMLS_STRING_SEP.join(semtype_category))
                    if semtype_priority > concept_semtype_priority:
                        concept_semtype_priority = semtype_priority
                concept_semtypes = '%'.join(concept_semtypes)

                # some concepts have their priority and score changed
                if concept_candidate_preferred in MMQUERY_SCORE_MOD:
                    if MMQUERY_SCORE_MOD[concept_candidate_preferred][0]:
                        concept_semtype_priority = MMQUERY_SCORE_MOD[concept_candidate_preferred][0]
                    if MMQUERY_SCORE_MOD[concept_candidate_preferred][1]:
                        concept_score = MMQUERY_SCORE_MOD[concept_candidate_preferred][1]

                # get the current candidate record fields and update the best candidates
                current_candidate_record = [concept_score, concept_cui, concept_candidate_preferred,
                                            concept_semtypes, concept_candidate_matched]

                for allowed_concept_cat in ALLOWED_CONCEPT_CATS:
                    if allowed_concept_cat in current_candidate_record[3]:
                        if current_candidate_record[1] not in BLACKLISTED_CUIS:
                            candidates.append((current_candidate_record,concept_semtype_priority))
                        all_concepts.append(concept_candidate_preferred)
                        break
                    else:
                        continue


    best_candidate_rating = [0, 0]
    best_candidate_record = [MMQUERY_CSV_DEFAULT_FIELD]*5

    for candidate in candidates:
        best_candidate_record, best_candidate_rating = update_record(best_candidate_rating, [candidate[1], candidate[0][0]],
                                                                         best_candidate_record, candidate[0])

    if ((best_candidate_record[1]=="C2081627") or (best_candidate_record[2]=="plan: surgery (treatment)")) :
        best_candidate_record[1] = "C0543467"
        best_candidate_record[2] = "Operative Surgical Procedures"
        best_candidate_record[3] = "PROC|Procedures|T061|Therapeutic or Preventive Procedure"
        best_candidate_record[4] = "SURGERY"

    return best_candidate_record, all_concepts


# Timeout after 10 seconds for single request strings
@timeout(10)
def invoke_mm1(cmd_task):
    return subprocess.check_output(cmd_task, shell=True).split(MMQUERY_XML_HEADER)[1]

def mm_problematic_batch1(batch_strings,umls_cat_abbrv,umls_cat_group):

    metamap_results = []
    for request_string in batch_strings:

        print("Processing single string")
        cmd_task = MMQUERY_CMD(request_string)
        try:
            out_xml = invoke_mm1(cmd_task)
        except:
            logging.warning('Metamap fails on string: ' + str(request_string))
            out_xml = None

        mapped_primary, all_concepts = get_best_record(out_xml,request_string,umls_cat_abbrv,umls_cat_group)
        #Keep unique candidates
        all_concepts = list(set(all_concepts))
        if (len(all_concepts) > 0):
            all_concepts = (" || ").join(all_concepts)
        else:
            all_concepts = MMQUERY_CSV_DEFAULT_FIELD
        metamap_results.append((mapped_primary,all_concepts,request_string))

    return metamap_results


# Timeout after 30 seconds for batches of 10
@timeout(30)
def invoke_mm10(cmd_task):
    return subprocess.check_output(cmd_task, shell=True).split(MMQUERY_XML_HEADER)[1:]

def mm_problematic_batch10(batch_strings,umls_cat_abbrv,umls_cat_group):

    metamap_results = []

    for i in range(0, len(batch_strings), 10):

        print("Processing in batches of 10")
        request_strings = batch_strings[i:i+10]
        query = '\n '.join(request_strings)
        cmd_task = MMQUERY_CMD(query)

        try:
            batch_xmls = invoke_mm10(cmd_task)
        except:
            batch_xmls  = []
            logging.warning('Metamap exception 10')

        if len(batch_xmls) != len(request_strings):
            logging.warning('Metamap error 10')
            metamap_results.extend(mm_problematic_batch1(request_strings,umls_cat_abbrv,umls_cat_group))

        else:

            for idx,xml in enumerate(batch_xmls):

                mapped_primary, all_concepts = get_best_record(xml, request_strings[idx],umls_cat_abbrv,umls_cat_group)

                #Keep unique candidates
                all_concepts = list(set(all_concepts))
                if (len(all_concepts) > 0):
                    all_concepts = (" || ").join(all_concepts)
                else:
                    all_concepts = MMQUERY_CSV_DEFAULT_FIELD
                metamap_results.append((mapped_primary,all_concepts,request_strings[idx]))

    return metamap_results


# Timeout after 200 seconds for batches of 100
@timeout(200)
def invoke_mm100(cmd_task):
    return subprocess.check_output(cmd_task, shell=True).split(MMQUERY_XML_HEADER)[1:]

def mm_problematic_batch100(batch_strings,umls_cat_abbrv,umls_cat_group):


    metamap_results = []

    for i in range(0, len(batch_strings), 100):

        print("Processing in batches of 100")
        request_strings = batch_strings[i:i+100]
        query = '\n '.join(request_strings)
        cmd_task = MMQUERY_CMD(query)
        try:
            batch_xmls = invoke_mm100(cmd_task)
        except:
            batch_xmls = []
            logging.warning('Metamap exception 100')
            logging.warning(str(len(batch_xmls)))

        if len(batch_xmls) != len(request_strings):
            logging.warning('Metamap error 100')
            metamap_results.extend(mm_problematic_batch10(request_strings,umls_cat_abbrv,umls_cat_group))

        else:

            for idx,xml in enumerate(batch_xmls):

                mapped_primary, all_concepts = get_best_record(xml, request_strings[idx],umls_cat_abbrv,umls_cat_group)
                #Keep unique candidates
                all_concepts = list(set(all_concepts))
                if (len(all_concepts) > 0):
                    all_concepts = (" || ").join(all_concepts)
                else:
                    all_concepts = MMQUERY_CSV_DEFAULT_FIELD
                metamap_results.append((mapped_primary,all_concepts,request_strings[idx]))

    return metamap_results




# Timeout after 1500 seconds for full batches
@timeout(1500)
def invoke_mm(cmd_task):
    return subprocess.check_output(cmd_task, shell=True).split(MMQUERY_XML_HEADER)[1:]

if __name__ == '__main__':

    from querying_mappings import sym2text
    from global_chem_mappings import global_chem2text
    from global_term_mappings import global_text2text


    # memoize overrides the MetaMap dictionary containing all query results
    # MMQUERY_MM_MEMOIZE = False
    MMQUERY_MM_MEMOIZE = True

    MMQUERY_INTERVENTION_INDEX = 10
    MMQUERY_INTERVENTION_REQUEST = 11
    MMQUERY_INTERVENTION_CONCEPT = 12
    MMQUERY_INTERVENTION_CUI = 13
    MMQUERY_INTERVENTION_CATEGORY = 14
    MMQUERY_INTERVENTION_ALL_CONCEPTS = 15



    # set of characters that cannot appear within a MetaMap query
    MMQUERY_FORBIDDEN_SYMBOLS = set(['.',':','"','`','´','>','<','+','\\'])

    # instance of an Spacy lemmatizer
    MMQUERY_NLP = spacy.load('en_core_web_sm', disable=['tagger', 'parser', 'ner', 'textcat'])


    # header of the XML files produced by MetaMap
    MMQUERY_XML_HEADER = '<?xml version="1.0" encoding="UTF-8"?>'

    # input file name
    input_file = sys.argv[1]
    # output file name
    output_file = sys.argv[2]

    # location of the pre-computed MetaMap dictionary
    MMQUERY_METAMAP_DICT = sys.argv[3]


    # read UMLS categories
    umls_cat_abbrv, umls_cat_group = read_umls_catcodes()

    # size of batch requests to send to MetaMap
    REQUEST_SIZE = 1000
    SAVE_PER_NUM_BATCHES = 1

    # concept strings in each request within the batch
    batch_request_strings = []

    # modified input rows to output
    output_rows = []

    # whether it is necessary to issue a MetaMap request for each of the input rows
    output_rows_dict_mapped = []

    # whether the current row is the header
    header_row = True

    # index of the current row
    row_index = 0

    # index of the last saved row
    previous_row_index = 1

    # Increase field size limit to max
    # csv.field_size_limit(sys.maxsize)
    # store the input file in memory
    csv_in_rows = list(csv.reader(io.open(input_file, mode='r', encoding='utf-8'), delimiter = MMQUERY_CSV_DELIMITER))

    # number of rows in the input file
    row_count = len(csv_in_rows)

    # load the preloaded metamap dictionary
    if not MMQUERY_MM_MEMOIZE:
        with open(MMQUERY_METAMAP_DICT, 'rb') as handle:
            premm_dictionary = pickle.load(handle)
    else:
        # Manually assigned records
        premm_dictionary = dict()
        # Read manually assigned concepts .tsv file for interventions
        conditions_remap_file = "utils/CT_interventions_manual_remaps.tsv"
        with open(conditions_remap_file, 'r') as my_file:
            reader = csv.reader(my_file, delimiter='\t', quoting=csv.QUOTE_NONE, quotechar='')

            for idx, row in enumerate(reader):

                if idx == 0:
                     # Map column name to their indices
                    fieldIndex = dict()
                    for pair in list(enumerate(row)):
                        fieldIndex[pair[1]] = pair[0]
                else:
                    remapped_request = row[fieldIndex["Intervention_request"]]
                    remapped_cui = row[fieldIndex["Intervention_Cui"]]
                    remapped_concept = row[fieldIndex["Intervention_Concept"]]
                    remapped_category = row[fieldIndex["Intervention_CAT"]]
                    ignore_flag = row[fieldIndex["Ignore"]]


                    if remapped_cui == 0 or remapped_cui == "0" or remapped_cui == '0' or remapped_cui == None or remapped_cui.strip() == "" or ignore_flag == 1 or ignore_flag == '1' or ignore_flag =="1":
                        premm_dictionary[remapped_request] = ([MMQUERY_CSV_DEFAULT_FIELD]*5 + [1], MMQUERY_CSV_DEFAULT_FIELD)
                    else:
                        premm_dictionary[remapped_request] = ([1000,remapped_cui, remapped_concept, remapped_category,remapped_request] + [1], MMQUERY_CSV_DEFAULT_FIELD)

            premm_dictionary[0] = ([MMQUERY_CSV_DEFAULT_FIELD]*5 + [1], MMQUERY_CSV_DEFAULT_FIELD)
            premm_dictionary['0'] = ([MMQUERY_CSV_DEFAULT_FIELD]*5 + [1], MMQUERY_CSV_DEFAULT_FIELD)
            premm_dictionary["0"] = ([MMQUERY_CSV_DEFAULT_FIELD]*5 + [1], MMQUERY_CSV_DEFAULT_FIELD)

        # for k,v in premm_dictionary.items():
        #     print(k)
        #     print(v)
        # print(len(premm_dictionary))
        # sys.exit()
    # Measure per-batch processing time
    start = time.time()
    start2 = time.time()
    # read information from the CSV file and process each row
    mm_batch_num = 0

    for row in csv_in_rows:
        # add a new field to the input rows
        # this field indicates what region was used by MetaMap to match the input string
        if header_row:
            header_row = False
            output_rows.append(row)
            output_rows_dict_mapped.append([True])
            row_index += 1
            continue
        else:
            output_rows.append(row)
            output_rows_dict_mapped.append([])


        term_display_nohyp = map_global_terms(clean_entity_string(substract_hyphens(output_rows[-1][MMQUERY_INTERVENTION_INDEX].lower())))
        request_string = clean_query_string(heuristic_fixes(term_display_nohyp))

        if request_string in premm_dictionary:
            [score, cui, preferred, semtypes, matched_candidate, count] = premm_dictionary[request_string][0]
            premm_dictionary[request_string][0][5] = count+1
            all_concepts =  premm_dictionary[request_string][1]
            output_rows[-1][MMQUERY_INTERVENTION_REQUEST] = request_string
            output_rows[-1][MMQUERY_INTERVENTION_CONCEPT] = preferred
            output_rows[-1][MMQUERY_INTERVENTION_CUI] = cui
            output_rows[-1][MMQUERY_INTERVENTION_CATEGORY] = semtypes
            output_rows[-1][MMQUERY_INTERVENTION_ALL_CONCEPTS] = all_concepts.replace(MMQUERY_CSV_DELIMITER, ',')
            output_rows_dict_mapped[-1].append(True)


        # otherwise we need to query MetaMap, but we will launch queries in batches for speed
        else:

            # add request strings to the current batch
            if ((request_string.split()) and (request_string.strip() not in string.punctuation)):

                batch_request_strings.append(request_string)
                output_rows_dict_mapped[-1].append(False)

            # avoid querying MetaMap for empty strings or punctuation chars, and consider them completed
            else:

                output_rows[-1][MMQUERY_INTERVENTION_REQUEST] = request_string
                output_rows[-1][MMQUERY_INTERVENTION_CONCEPT] = MMQUERY_CSV_DEFAULT_FIELD
                output_rows[-1][MMQUERY_INTERVENTION_CUI] = MMQUERY_CSV_DEFAULT_FIELD
                output_rows[-1][MMQUERY_INTERVENTION_CATEGORY] = MMQUERY_CSV_DEFAULT_FIELD
                output_rows[-1][MMQUERY_INTERVENTION_ALL_CONCEPTS] = MMQUERY_CSV_DEFAULT_FIELD
                output_rows_dict_mapped[-1].append(True)


        # if the current batch is sufficiently large, run a MetaMap query
        if len(batch_request_strings) >= REQUEST_SIZE or (row_index + 1) == row_count:
            # construct query by separating terms with newlines
            query = '\n '.join(batch_request_strings)
            # execute query
            cmd_task = MMQUERY_CMD(query)
            try:
                batch_xmls = invoke_mm(cmd_task)
            except:
                batch_xmls = []
                logging.warning('Metamap error in rows: [' + str(row_index) + ' , ' + str(row_index-int(REQUEST_SIZE/2)-1) + ']')
                print('\x1b[0;30;41m' + 'Metamap error in rows: [', row_index, ',', row_index-int(REQUEST_SIZE/2)-1, ']' + '\x1b[0m')

            metamap_results_primary = []

            if len(batch_xmls) == len(batch_request_strings):

                xml_index = 0
                for xml_data in batch_xmls:
                    mapped_primary, all_concepts = get_best_record(xml_data, batch_request_strings[xml_index],
                                                     umls_cat_abbrv,umls_cat_group)

                    # Keep only unique concepts
                    all_concepts = list(set(all_concepts))
                    if (len(all_concepts) > 0):
                        all_concepts = (" || ").join(all_concepts)
                    else:
                        all_concepts = MMQUERY_CSV_DEFAULT_FIELD

                    metamap_results_primary.append((mapped_primary,all_concepts,batch_request_strings[xml_index]))


                    if MMQUERY_MM_MEMOIZE:
                        premm_dictionary[batch_request_strings[xml_index]] = (mapped_primary + [1], all_concepts)
                    xml_index += 1

            else:
                # Do not cache when a batch fails
                print('\x1b[0;30;41m' + 'Metamap2 error in rows: [', row_index, ',', row_index-int(REQUEST_SIZE/2)-1, ']' + '\x1b[0m')
                print(len(batch_request_strings),len(metamap_results_primary),len(batch_xmls))
                logging.warning('Metamap2 error in rows: [' + str(row_index) + ' , '+ str(row_index-int(REQUEST_SIZE/2)-1) + ']')
                metamap_results_primary = mm_problematic_batch100(batch_request_strings,umls_cat_abbrv,umls_cat_group)


            # write the output CSV file
            out_rows_index = len(output_rows) - 1

            for i in reversed(range(len(metamap_results_primary))):
                # obtain the terms to write
                [score, cui, preferred, semtypes, matched_candidate] = metamap_results_primary[i][0]
                all_concepts = metamap_results_primary[i][1]
                r_string = metamap_results_primary[i][2]


                # obtain the index where the terms should be written
                while output_rows_dict_mapped[out_rows_index][0]:
                    out_rows_index -= 1

                output_rows[out_rows_index][MMQUERY_INTERVENTION_REQUEST] = r_string
                output_rows[out_rows_index][MMQUERY_INTERVENTION_CONCEPT] = preferred
                output_rows[out_rows_index][MMQUERY_INTERVENTION_CUI] = cui
                output_rows[out_rows_index][MMQUERY_INTERVENTION_CATEGORY] = semtypes
                output_rows[out_rows_index][MMQUERY_INTERVENTION_ALL_CONCEPTS] = all_concepts.replace(MMQUERY_CSV_DELIMITER, ',')
                output_rows_dict_mapped[out_rows_index][0] = True

            #reset current batch
            batch_request_strings = []
            print("Finish batch  in %.2fs" %  (time.time() - start))
            mm_batch_num += 1
            if not mm_batch_num % SAVE_PER_NUM_BATCHES:
                with open(output_file, 'w') as f:
                    w = csv.writer(f, delimiter=MMQUERY_CSV_DELIMITER, quoting = csv.QUOTE_NONE, quotechar='', escapechar='\\')
                    w.writerows(output_rows)
                print("Processed rows: %i Total rows: %i  in %.2fs" % (row_index,row_count,(time.time() - start2)))
                previous_row_index = row_index + 1

            start = time.time()

        row_index += 1

    if previous_row_index!=row_index:
        with open(output_file, 'w') as f:
            w = csv.writer(f, delimiter=MMQUERY_CSV_DELIMITER, quoting = csv.QUOTE_NONE, quotechar='',escapechar='\\')
            w.writerows(output_rows)
        print("Processed rows: %i Total rows: %i  in %.2fs" % (row_index,row_count,(time.time() - start2)))


    # Rerun failed requests
    print("Resend to metamap all failed request strings")
    failed_requests_mappings = dict()
    if len(FAILED_REQUESTS) > 0:
        logging.warning('Rerun metamap for failed request strings')
        for req, _ in FAILED_REQUESTS.items():
            cmd_task = MMQUERY_CMD(req)
            try:
                out_xml = invoke_mm1(cmd_task)
            except:
                logging.warning('Metamap fails on string: ' + str(req))
                out_xml = None

            mapped_primary, all_concepts = get_best_record(out_xml,request_string,umls_cat_abbrv,umls_cat_group)

            all_concepts = list(set(all_concepts))
            if (len(all_concepts) > 0):
                all_concepts = (" || ").join(all_concepts)
            else:
                all_concepts = MMQUERY_CSV_DEFAULT_FIELD
            # print(req)
            # print(mapped_primary)
            # print("=====")
            if mapped_primary[1] != MMQUERY_CSV_DEFAULT_FIELD:
                failed_requests_mappings[req] = [mapped_primary, all_concepts]

    # Update initially failed requests on the output file
    if len(failed_requests_mappings) > 0:
        count_updates = 0
        for out_index in range(len(output_rows)):
            if out_index == 0:
                continue
            else:
                if output_rows[out_index][MMQUERY_INTERVENTION_REQUEST] in failed_requests_mappings:
                    count_updates += 1
                    [score, cui, preferred, semtypes, matched_candidate] = failed_requests_mappings[output_rows[out_index][MMQUERY_INTERVENTION_REQUEST]][0]
                    all_concepts = failed_requests_mappings[output_rows[out_index][MMQUERY_INTERVENTION_REQUEST]][1]
                    output_rows[out_index][MMQUERY_INTERVENTION_CONCEPT] = preferred
                    output_rows[out_index][MMQUERY_INTERVENTION_CUI] = cui
                    output_rows[out_index][MMQUERY_INTERVENTION_CATEGORY] = semtypes
                    output_rows[out_index][MMQUERY_INTERVENTION_ALL_CONCEPTS] = all_concepts.replace(MMQUERY_CSV_DELIMITER, ',')
                    premm_dictionary[output_rows[out_index][MMQUERY_INTERVENTION_REQUEST]] = (failed_requests_mappings[output_rows[out_index][MMQUERY_INTERVENTION_REQUEST]][0] + [1], all_concepts)


        print("Total updates for initially failed requests: "+str(count_updates))
        print("Re-write output file ...")
        with open(output_file, 'w') as f:
            w = csv.writer(f, delimiter=MMQUERY_CSV_DELIMITER, quoting = csv.QUOTE_NONE, quotechar='',escapechar='\\')
            w.writerows(output_rows)
        print("Done")



    print("Write metamap cache ...")
    #write the precomputed metamap
    if MMQUERY_MM_MEMOIZE:
        with open(MMQUERY_METAMAP_DICT, 'wb') as handle:
            pickle.dump(premm_dictionary, handle, protocol=pickle.HIGHEST_PROTOCOL)


    print('Done')
