import csv
import sys

from openpyxl import Workbook
from copy import deepcopy

csv.field_size_limit(58000000)


umls_2023AB_file = sys.argv[1]
umls_2025AB_file = sys.argv[2]
out_xlsx = sys.argv[3]


umls_2023AB_conditions = {}
umls_2023AB_interventions = {}
umls_2025AB_conditions = {}
umls_2025AB_interventions = {}
conditions_different_count = 0
interventions_different_count = 0
evidence_2023AB = {}
trials_2023AB = {}
evidence_2025AB = {}
trials_2025AB = {}

print("Reading 2023AB output ...")
with open(umls_2023AB_file, 'r') as my_file:
    reader = csv.reader(my_file, delimiter='\t', quoting=csv.QUOTE_NONE, quotechar='', escapechar='\\')

    for idx, row in enumerate(reader):

        if idx == 0:

            # Map column name to their indices
            fieldIndex = dict()
            for pair in list(enumerate(row)):
                fieldIndex[pair[1]] = pair[0]
        else:
            filename = row[fieldIndex["filename"]].split("/")[-1].split(".")[0]
            cond_request = row[fieldIndex["cond_mm_request"]]
            cond_concept = row[fieldIndex["condition_concept"]]
            cond_cui = row[fieldIndex["condition_cui"]]
            cond_cat = row[fieldIndex["condition_categories"]]
            cond_all_mm = row[fieldIndex["condition_all_mm"]]

            if cond_request not in umls_2023AB_conditions:
                umls_2023AB_conditions[cond_request] = [cond_concept, cond_cui, cond_cat, cond_all_mm, [filename]]
            else:
                if cond_cui != umls_2023AB_conditions[cond_request][1]:
                    print("Error! Different condition concept for the same request!!!")
                    sys.exit()
                umls_2023AB_conditions[cond_request][-1].append(filename)

            inter_request = row[fieldIndex["intervention_mm_request"]]
            inter_concept = row[fieldIndex["intervention_concept"]]
            inter_cui = row[fieldIndex["intervention_cui"]]
            inter_cat = row[fieldIndex["intervention_categories"]]
            inter_all_mm = row[fieldIndex["intervention_all_mm"]]


            if inter_request not in umls_2023AB_interventions:
                umls_2023AB_interventions[inter_request] = [inter_concept, inter_cui, inter_cat,
                                                            inter_all_mm, [filename]]
            else:
                if inter_cui != umls_2023AB_interventions[inter_request][1]:
                    print("Error! Different intervention concept for the same request!!!")
                    sys.exit()
                umls_2023AB_interventions[inter_request][-1].append(filename)
            
            if inter_cui == '0' and cond_cui == '0':
                evidence_2023AB['|'.join([filename,cond_request.strip().lower(),inter_request.strip().lower()])] = 1
            elif cond_cui == '0':
                evidence_2023AB['|'.join([filename,cond_request.strip().lower(),inter_cui])] = 1
            elif inter_cui == '0':
                evidence_2023AB['|'.join([filename,cond_cui,inter_request.strip().lower()])] = 1
            else:
                evidence_2023AB['|'.join([filename,cond_cui,inter_cui])] = 1

            trials_2023AB[filename] = 1



print("Total evidence rows for 2023AB: ", len(evidence_2023AB))
print("Sample of evidence triplets: ", list(evidence_2023AB.keys())[:10])
print("Total trials: ", len(trials_2023AB))
print("Done")

print("--------------------------------")
print("Reading 2025AB output ...")
with open(umls_2025AB_file, 'r') as my_file:
    reader = csv.reader(my_file, delimiter='\t', quoting=csv.QUOTE_NONE, quotechar='', escapechar='\\')

    for idx, row in enumerate(reader):

        if idx == 0:

            # Map column name to their indices
            fieldIndex = dict()
            for pair in list(enumerate(row)):
                fieldIndex[pair[1]] = pair[0]
        else:
            filename = row[fieldIndex["filename"]].split("/")[-1].split(".")[0]
            cond_request = row[fieldIndex["cond_mm_request"]]
            cond_concept = row[fieldIndex["condition_concept"]]
            cond_cui = row[fieldIndex["condition_cui"]]
            cond_cat = row[fieldIndex["condition_categories"]]
            cond_all_mm = row[fieldIndex["condition_all_mm"]]

            if cond_request not in umls_2025AB_conditions:
                umls_2025AB_conditions[cond_request] = [cond_concept, cond_cui, cond_cat, cond_all_mm, [filename]]
            else:
                if cond_cui != umls_2025AB_conditions[cond_request][1]:
                    print("Error! Different condition concept for the same request!!!")
                    sys.exit()
                umls_2025AB_conditions[cond_request][-1].append(filename)

            inter_request = row[fieldIndex["intervention_mm_request"]]
            inter_concept = row[fieldIndex["intervention_concept"]]
            inter_cui = row[fieldIndex["intervention_cui"]]
            inter_cat = row[fieldIndex["intervention_categories"]]
            inter_all_mm = row[fieldIndex["intervention_all_mm"]]

            if inter_request not in umls_2025AB_interventions:
                umls_2025AB_interventions[inter_request] = [inter_concept, inter_cui, inter_cat,
                                                            inter_all_mm, [filename]]
            else:
                if inter_cui != umls_2025AB_interventions[inter_request][1]:
                    print("Error! Different intervention concept for the same request!!!")
                    sys.exit()
                umls_2025AB_interventions[inter_request][-1].append(filename)
            
            if inter_cui == '0' and cond_cui == '0':
                evidence_2025AB['|'.join([filename,cond_request.strip().lower(),inter_request.strip().lower()])] = 1
            elif cond_cui == '0':
                evidence_2025AB['|'.join([filename,cond_request.strip().lower(),inter_cui])] = 1
            elif inter_cui == '0':
                evidence_2025AB['|'.join([filename,cond_cui,inter_request.strip().lower()])] = 1
            else:
                evidence_2025AB['|'.join([filename,cond_cui,inter_cui])] = 1
            trials_2025AB[filename] = 1

print("Total evidence rows for 2025AB: ", len(evidence_2025AB))
print("Sample of evidence triplets: ", list(evidence_2025AB.keys())[:10])
print("Total trials: ", len(trials_2025AB))
print("Done")
print("--------------------------------")

wb = Workbook()
ws1 = wb.create_sheet("Conditions_changed")
ws1.title = "Conditions_changed"
ws2 = wb.create_sheet("Interventions_changed")
ws2.title = "Interventions_changed"
del wb["Sheet"]

# TODO add sheets for new conditions and interventions

conditions_sheet = wb['Conditions_changed']
conditions_sheet["A1"] = "Condition request string"
conditions_sheet["B1"] = "CONCEPT_2023AB"
conditions_sheet["C1"] = "CONCEPT_2025AB"
conditions_sheet["D1"] = "CUI_2023AB"
conditions_sheet["E1"] = "CUI_2025AB"
conditions_sheet["F1"] = "CAT_2023AB"
conditions_sheet["G1"] = "CAT_2025AB"
conditions_sheet["H1"] = "ALL_Concepts_2023AB"
conditions_sheet["I1"] = "ALL_Concepts_2025AB"
conditions_sheet["J1"] = "Frequency"
conditions_sheet["K1"] = "NCTs_sample"

row_index = 1
for k, v in umls_2023AB_conditions.items():
    if v[1] != umls_2025AB_conditions[k][1]:
        conditions_different_count += 1
        v_2023AB = deepcopy(v)
        for v_idx in range(len(v_2023AB)-1):
            if v_2023AB[v_idx] == '0' or v_2023AB[v_idx] == "0" or v_2023AB[v_idx].strip() == "":
                v_2023AB[v_idx] = 0
        for v_idx in range(len(umls_2025AB_conditions[k]) - 1):
            if umls_2025AB_conditions[k][v_idx] == '0' or umls_2025AB_conditions[k][v_idx] == "0" or \
                    umls_2025AB_conditions[k][v_idx].strip() == "":
                umls_2025AB_conditions[k][v_idx] = 0
        row_index += 1
        conditions_sheet['A' + str(row_index)] = k
        conditions_sheet['B' + str(row_index)] = v_2023AB[0]
        conditions_sheet['C' + str(row_index)] = umls_2025AB_conditions[k][0]
        conditions_sheet['D' + str(row_index)] = v_2023AB[1]
        conditions_sheet['E' + str(row_index)] = umls_2025AB_conditions[k][1]
        conditions_sheet['F' + str(row_index)] = v_2023AB[2]
        conditions_sheet['G' + str(row_index)] = umls_2025AB_conditions[k][2]
        conditions_sheet['H' + str(row_index)] = v_2023AB[3]
        conditions_sheet['I' + str(row_index)] = umls_2025AB_conditions[k][3]
        conditions_sheet['J' + str(row_index)] = len(list(set(v_2023AB[4])))
        conditions_sheet['K' + str(row_index)] = "|".join(list(set(v_2023AB[4]))[:10])

# Sort conditions sheet by frequency (column J) in descending order
data_rows = []
for row in conditions_sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is not None:  # Skip empty rows
        data_rows.append(row)

# Sort by frequency (column J, index 9) in descending order
data_rows.sort(key=lambda x: x[9] if x[9] is not None else 0, reverse=True)

# Clear existing data (keep headers)
for row in conditions_sheet.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

# Write sorted data back
for idx, row_data in enumerate(data_rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        conditions_sheet.cell(row=idx, column=col_idx, value=value)

interventions_sheet = wb['Interventions_changed']
interventions_sheet["A1"] = "Intervention request string"
interventions_sheet["B1"] = "CONCEPT_2023AB"
interventions_sheet["C1"] = "CONCEPT_2025AB"
interventions_sheet["D1"] = "CUI_2023AB"
interventions_sheet["E1"] = "CUI_2025AB"
interventions_sheet["F1"] = "CAT_2023AB"
interventions_sheet["G1"] = "CAT_2025AB"
interventions_sheet["H1"] = "ALL_Concepts_2023AB"
interventions_sheet["I1"] = "ALL_Concepts_2025AB"
interventions_sheet["J1"] = "Frequency"
interventions_sheet["K1"] = "NCTs_sample"

row_index = 1
for k, v in umls_2023AB_interventions.items():
    if v[1] != umls_2025AB_interventions[k][1]:
        interventions_different_count += 1
        v_2023AB = deepcopy(v)
        for v_idx in range(len(v_2023AB)-1):
            if v_2023AB[v_idx] == '0' or v_2023AB[v_idx] == "0" or v_2023AB[v_idx].strip() == "":
                v_2023AB[v_idx] = 0
        for v_idx in range(len(umls_2023AB_interventions[k]) - 1):
            if umls_2023AB_interventions[k][v_idx] == '0' or umls_2023AB_interventions[k][v_idx] == "0" or \
                    umls_2023AB_interventions[k][v_idx].strip() == "":
                umls_2023AB_interventions[k][v_idx] = 0
        row_index += 1
        interventions_sheet['A' + str(row_index)] = k
        interventions_sheet['B' + str(row_index)] = v_2023AB[0]
        interventions_sheet['C' + str(row_index)] = umls_2025AB_interventions[k][0]
        interventions_sheet['D' + str(row_index)] = v_2023AB[1]
        interventions_sheet['E' + str(row_index)] = umls_2025AB_interventions[k][1]
        interventions_sheet['F' + str(row_index)] = v_2023AB[2]
        interventions_sheet['G' + str(row_index)] = umls_2025AB_interventions[k][2]
        interventions_sheet['H' + str(row_index)] = v_2023AB[3]
        interventions_sheet['I' + str(row_index)] = umls_2025AB_interventions[k][3]
        interventions_sheet['J' + str(row_index)] = len(list(set(v_2023AB[4])))
        interventions_sheet['K' + str(row_index)] = "|".join(list(set(v_2023AB[4]))[:10])

# Sort interventions sheet by frequency (column J) in descending order
data_rows = []
for row in interventions_sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is not None:  # Skip empty rows
        data_rows.append(row)

# Sort by frequency (column J, index 9) in descending order
data_rows.sort(key=lambda x: x[9] if x[9] is not None else 0, reverse=True)

# Clear existing data (keep headers)
for row in interventions_sheet.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

# Write sorted data back
for idx, row_data in enumerate(data_rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        interventions_sheet.cell(row=idx, column=col_idx, value=value)
wb.save(filename=out_xlsx)

overlap_evidence_count = 0
new_changed_evidence_count = 0
for k,v in evidence_2025AB.items():
    if k in evidence_2023AB:
        overlap_evidence_count += 1
    else:
        new_changed_evidence_count += 1

print(f"Overlap evidence count: {overlap_evidence_count}")
print(f"New/changed evidence count: {new_changed_evidence_count}")

conditions_different_count_percentage = (conditions_different_count / len(umls_2023AB_conditions)) * 100
interventions_different_count_percentage = (interventions_different_count / len(umls_2023AB_interventions)) * 100
print(f"Conditions different count: {conditions_different_count}")
print(f"Interventions different count: {interventions_different_count}")
print(f"Conditions different count percentage: {conditions_different_count_percentage}%")
print(f"Interventions different count percentage: {interventions_different_count_percentage}%")

print(f"Total conditions: {len(umls_2023AB_conditions)}")
print(f"Total interventions: {len(umls_2023AB_interventions)}")
