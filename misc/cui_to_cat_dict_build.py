import pickle
import sys

from openpyxl import load_workbook

mrsty_file = sys.argv[1]
umls_cat_xlsx = sys.argv[2]
out_pickle_name = sys.argv[3]


def read_umls_catcodes(categories_xlsx):
    """
    Read UMLS categories and creates a dictionary with T_codes as keys
    and categories as values e.g: 'T192': CHEM|Chemicals & Drugs|T192|Receptor
    """

    t_code2cat = {}
    wb = load_workbook(filename=categories_xlsx)
    sheet = wb['sem_groups']
    start_row = 2
    end_row = sheet.max_row + 1
    for row_index in range(start_row, end_row):
        row = sheet['A'+str(row_index)].value
        if len(row.split("|")) != 4:
            print("ERROR! Category entry with invalid format")
            print(row)
            sys.exit()
        t_code = row.split("|")[2]
        if t_code not in t_code2cat:
            t_code2cat[t_code] = row
        else:
            print("Duplicate entry")
            print(row)
            print(t_code2cat[t_code])

    return t_code2cat


t2cat = read_umls_catcodes(umls_cat_xlsx)
print(len(t2cat))

cui2cat = {}
mrsty = open(mrsty_file, 'r')
line_counter = 0

while True:
    line_counter += 1
    if not line_counter % 100000:
        print(line_counter)

    # Get next line from file
    line = mrsty.readline()
    if not line:
        break
    line_vals = line.split("|")
    cui = line_vals[0]
    T_code = line_vals[1]

    if cui not in cui2cat:
        cui2cat[cui] = [t2cat[T_code]]
    else:
        additional_cat = t2cat[T_code]
        if additional_cat not in cui2cat[cui]:
            cui2cat[cui].append(additional_cat)


print("Lines read: {}".format(line_counter))
mrsty.close()
print("cui_to_cat size: {}".format(len(cui2cat)))

missing_cuis_dict = {'C3650840':['T061'],
'C0650607': ['T116','T121'] ,
'C1169678': ['T200'],
'C0729543': ['T047'],
'C1720654': ['T122'],
'C0718005': ['T109', 'T121'],
'C1959833': ['T060'],
'C1320708': ['T061'],
'C0553968': ['T047'],
'C1828075': ['T047'],
'C2732254': ['T058'],
'C1293331': ['T061'],
'C1827140': ['T061'],
'C3853287': ['T116','T121'],
'C1301829': ['T047'],
'C2585437': ['T061'],
'C1289988': ['T074'],
'C1959921': ['T168'],
'C3472375': ['T191'],
'C2585119': ['T061'],
'C1293625': ['T061'],
'C1293232': ['T061'],
'C1273006': ['T122'],
'C1253707': ['T200'],
'C1443734': ['T121'],
'C1293440': ['T061']}

for k, v in missing_cuis_dict.items():
    cui2cat[k] = [t2cat[x] for x in v]


with open(out_pickle_name, "wb") as fh:
    pickle.dump(cui2cat, fh)

out_pickle_name_python2 = out_pickle_name.split(".")[0] + ".pkl2"

with open(out_pickle_name_python2, "wb") as fh:
    pickle.dump(cui2cat, fh, protocol=2)
