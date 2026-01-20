import csv
import sys

from openpyxl import Workbook

umls_2022AA_baseline_file = sys.argv[1]
umls_2022AA_iter1_file = sys.argv[2]
out_xlsx = sys.argv[3]

umls_2022AA_baseline_conditions = {}
umls_2022AA_baseline_interventions = {}
umls_2022AA_iter1_conditions = {}
umls_2022AA_iter1_interventions = {}

print("Reading 2022AA baseline output ...")
with open(umls_2022AA_baseline_file, 'r') as my_file:
    reader = csv.reader(my_file, delimiter='\t', quoting=csv.QUOTE_NONE, quotechar='', escapechar='\\')

    for idx, row in enumerate(reader):

        if idx == 0:

            # Map column name to their indices
            fieldIndex = dict()
            for pair in list(enumerate(row)):
                fieldIndex[pair[1]] = pair[0]
        else:
            filename = row[fieldIndex["filename"]].split("/")[-1].split(".")[0]
            cond_request = row[fieldIndex["cond_mm_request"]]
            cond_concept = row[fieldIndex["condition_concept"]]
            cond_cui = row[fieldIndex["condition_cui"]]
            cond_cat = row[fieldIndex["condition_categories"]]
            cond_all_mm = row[fieldIndex["condition_all_mm"]]

            if cond_request not in umls_2022AA_baseline_conditions:
                umls_2022AA_baseline_conditions[cond_request] = [cond_concept, cond_cui, cond_cat, cond_all_mm, [filename]]
            else:
                if cond_cui != umls_2022AA_baseline_conditions[cond_request][1]:
                    print("Error! Different condition concept for the same request!!!")
                    sys.exit()
                umls_2022AA_baseline_conditions[cond_request][-1].append(filename)

            inter_request = row[fieldIndex["intervention_mm_request"]]
            inter_concept = row[fieldIndex["intervention_concept"]]
            inter_cui = row[fieldIndex["intervention_cui"]]
            inter_cat = row[fieldIndex["intervention_categories"]]
            inter_all_mm = row[fieldIndex["intervention_all_mm"]]

            if inter_request not in umls_2022AA_baseline_interventions:
                umls_2022AA_baseline_interventions[inter_request] = [inter_concept, inter_cui, inter_cat,
                                                                     inter_all_mm, [filename]]
            else:
                if inter_cui != umls_2022AA_baseline_interventions[inter_request][1]:
                    print("Error! Different intervention concept for the same request!!!")
                    sys.exit()
                umls_2022AA_baseline_interventions[inter_request][-1].append(filename)

print("Done")

print("Reading 2022AA iter 1 output ...")
with open(umls_2022AA_iter1_file, 'r') as my_file:
    reader = csv.reader(my_file, delimiter='\t', quoting=csv.QUOTE_NONE, quotechar='', escapechar='\\')

    for idx, row in enumerate(reader):

        if idx == 0:

            # Map column name to their indices
            fieldIndex = dict()
            for pair in list(enumerate(row)):
                fieldIndex[pair[1]] = pair[0]
        else:
            filename = row[fieldIndex["filename"]].split("/")[-1].split(".")[0]
            cond_request = row[fieldIndex["cond_mm_request"]]
            cond_concept = row[fieldIndex["condition_concept"]]
            cond_cui = row[fieldIndex["condition_cui"]]
            cond_cat = row[fieldIndex["condition_categories"]]
            cond_all_mm = row[fieldIndex["condition_all_mm"]]

            if cond_request not in umls_2022AA_iter1_conditions:
                umls_2022AA_iter1_conditions[cond_request] = [cond_concept, cond_cui, cond_cat, cond_all_mm, [filename]]
            else:
                if cond_cui != umls_2022AA_iter1_conditions[cond_request][1]:
                    print("Error! Different condition concept for the same request!!!")
                    sys.exit()
                umls_2022AA_iter1_conditions[cond_request][-1].append(filename)

            inter_request = row[fieldIndex["intervention_mm_request"]]
            inter_concept = row[fieldIndex["intervention_concept"]]
            inter_cui = row[fieldIndex["intervention_cui"]]
            inter_cat = row[fieldIndex["intervention_categories"]]
            inter_all_mm = row[fieldIndex["intervention_all_mm"]]

            if inter_request not in umls_2022AA_iter1_interventions:
                umls_2022AA_iter1_interventions[inter_request] = [inter_concept, inter_cui, inter_cat,
                                                                  inter_all_mm, [filename]]
            else:
                if inter_cui != umls_2022AA_iter1_interventions[inter_request][1]:
                    print("Error! Different intervention concept for the same request!!!")
                    sys.exit()
                umls_2022AA_iter1_interventions[inter_request][-1].append(filename)

print("Done")

wb = Workbook()
ws1 = wb.create_sheet("Conditions")
ws1.title = "Conditions"
ws2 = wb.create_sheet("Interventions")
ws2.title = "Interventions"
del wb["Sheet"]

conditions_sheet = wb['Conditions']
conditions_sheet["A1"] = "Condition request"
conditions_sheet["B1"] = "CONCEPT_2022AA_baseline"
conditions_sheet["C1"] = "CONCEPT_2022AA_iter1"
conditions_sheet["D1"] = "CUI_2022AA_baseline"
conditions_sheet["E1"] = "CUI_2022AA_iter1"
conditions_sheet["F1"] = "CAT_2022AA_baseline"
conditions_sheet["G1"] = "CAT_2022AA_iter1"
conditions_sheet["H1"] = "ALL_Concepts_2022AA_baseline"
conditions_sheet["I1"] = "ALL_Concepts_2022AA_iter1"
conditions_sheet["J1"] = "Frequency"
conditions_sheet["K1"] = "NCTs_sample"

row_index = 1
for k, v in umls_2022AA_baseline_conditions.items():
    if v[1] != umls_2022AA_iter1_conditions[k][1]:
        for v_idx in range(len(v)-1):
            if v[v_idx] == '0' or v[v_idx] == "0" or v[v_idx].strip() == "":
                v[v_idx] = 0
        for v_idx in range(len(umls_2022AA_iter1_conditions[k]) - 1):
            if umls_2022AA_iter1_conditions[k][v_idx] == '0' or umls_2022AA_iter1_conditions[k][v_idx] == "0" or \
                    umls_2022AA_iter1_conditions[k][v_idx].strip() == "":
                umls_2022AA_iter1_conditions[k][v_idx] = 0
        row_index += 1
        conditions_sheet['A' + str(row_index)] = k
        conditions_sheet['B' + str(row_index)] = v[0]
        conditions_sheet['C' + str(row_index)] = umls_2022AA_iter1_conditions[k][0]
        conditions_sheet['D' + str(row_index)] = v[1]
        conditions_sheet['E' + str(row_index)] = umls_2022AA_iter1_conditions[k][1]
        conditions_sheet['F' + str(row_index)] = v[2]
        conditions_sheet['G' + str(row_index)] = umls_2022AA_iter1_conditions[k][2]
        conditions_sheet['H' + str(row_index)] = v[3]
        conditions_sheet['I' + str(row_index)] = umls_2022AA_iter1_conditions[k][3]
        conditions_sheet['J' + str(row_index)] = len(list(set(v[4])))
        conditions_sheet['K' + str(row_index)] = "|".join(list(set(v[4]))[:10])

interventions_sheet = wb['Interventions']
interventions_sheet["A1"] = "Intervention request"
interventions_sheet["B1"] = "CONCEPT_2022AA_baseline"
interventions_sheet["C1"] = "CONCEPT_2022AA"
interventions_sheet["D1"] = "CUI_2022AA_baseline"
interventions_sheet["E1"] = "CUI_2022AA_iter1"
interventions_sheet["F1"] = "CAT_2022AA_baseline"
interventions_sheet["G1"] = "CAT_2022AA_iter1"
interventions_sheet["H1"] = "ALL_Concepts_2022AA_baseline"
interventions_sheet["I1"] = "ALL_Concepts_2022AA_iter1"
interventions_sheet["J1"] = "Frequency"
interventions_sheet["K1"] = "NCTs_sample"

row_index = 1
for k, v in umls_2022AA_baseline_interventions.items():
    if v[1] != umls_2022AA_iter1_interventions[k][1]:

        for v_idx in range(len(v)-1):
            if v[v_idx] == '0' or v[v_idx] == "0" or v[v_idx].strip() == "":
                v[v_idx] = 0
        for v_idx in range(len(umls_2022AA_iter1_interventions[k]) - 1):
            if umls_2022AA_iter1_interventions[k][v_idx] == '0' or umls_2022AA_iter1_interventions[k][v_idx] == "0" or \
                    umls_2022AA_iter1_interventions[k][v_idx].strip() == "":
                umls_2022AA_iter1_interventions[k][v_idx] = 0
        row_index += 1
        interventions_sheet['A' + str(row_index)] = k
        interventions_sheet['B' + str(row_index)] = v[0]
        interventions_sheet['C' + str(row_index)] = umls_2022AA_iter1_interventions[k][0]
        interventions_sheet['D' + str(row_index)] = v[1]
        interventions_sheet['E' + str(row_index)] = umls_2022AA_iter1_interventions[k][1]
        interventions_sheet['F' + str(row_index)] = v[2]
        interventions_sheet['G' + str(row_index)] = umls_2022AA_iter1_interventions[k][2]
        interventions_sheet['H' + str(row_index)] = v[3]
        interventions_sheet['I' + str(row_index)] = umls_2022AA_iter1_interventions[k][3]
        interventions_sheet['J' + str(row_index)] = len(list(set(v[4])))
        interventions_sheet['K' + str(row_index)] = "|".join(list(set(v[4]))[:10])
wb.save(filename=out_xlsx)
